from pathlib import Path
from typing import Dict, Any
from .base import BaseConverter
import openpyxl
from openpyxl.utils import get_column_letter
import xlrd

class ExcelConverter(BaseConverter):
    SUPPORTED_EXTENSIONS = ['.xlsx', '.xls']
    
    async def convert_to_markdown(self, file_path: Path) -> Dict[str, Any]:
        """Excel转Markdown"""
        try:
            ext = file_path.suffix.lower()
            if ext == '.xls':
                return self._convert_xls_to_markdown(file_path)
            
            wb = openpyxl.load_workbook(str(file_path), data_only=True)
            
            markdown_content = []
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                markdown_content.append(f"## {sheet_name}")
                markdown_content.append("")
                
                # 获取数据范围
                if ws.max_row > 0 and ws.max_column > 0:
                    # 创建表格
                    table = []
                    
                    # 获取表头
                    headers = []
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=1, column=col).value
                        headers.append(str(cell_value) if cell_value else "")
                    
                    table.append("| " + " | ".join(headers) + " |")
                    table.append("| " + " | ".join(["---"] * len(headers)) + " |")
                    
                    # 获取数据
                    for row in range(2, ws.max_row + 1):
                        row_data = []
                        for col in range(1, ws.max_column + 1):
                            cell_value = ws.cell(row=row, column=col).value
                            row_data.append(str(cell_value) if cell_value else "")
                        
                        if any(cell.strip() for cell in row_data):  # 跳过空行
                            table.append("| " + " | ".join(row_data) + " |")
                    
                    markdown_content.extend(table)
                    markdown_content.append("")
            
            return {
                "success": True,
                "content": "\n".join(markdown_content),
                "sheets": len(wb.sheetnames)
            }
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    def _convert_xls_to_markdown(self, file_path: Path) -> Dict[str, Any]:
        """处理旧版 .xls 格式"""
        wb = xlrd.open_workbook(str(file_path))
        markdown_content = []
        
        for sheet in wb.sheets():
            markdown_content.append(f"## {sheet.name}")
            markdown_content.append("")
            
            if sheet.nrows > 0 and sheet.ncols > 0:
                table = []
                headers = [str(sheet.cell_value(0, col)) if sheet.cell_value(0, col) else "" for col in range(sheet.ncols)]
                table.append("| " + " | ".join(headers) + " |")
                table.append("| " + " | ".join(["---"] * len(headers)) + " |")
                
                for row in range(1, sheet.nrows):
                    row_data = [str(sheet.cell_value(row, col)) if sheet.cell_value(row, col) else "" for col in range(sheet.ncols)]
                    if any(cell.strip() for cell in row_data):
                        table.append("| " + " | ".join(row_data) + " |")
                
                markdown_content.extend(table)
                markdown_content.append("")
        
        return {
            "success": True,
            "content": "\n".join(markdown_content),
            "sheets": wb.nsheets
        }
    
    async def convert_from_markdown(self, md_content: str, output_path: Path) -> Path:
        """Markdown转Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        lines = md_content.split('\n')
        in_table = False
        table_data = []
        headers = []
        
        for line in lines:
            line = line.strip()
            
            # 检测表格开始
            if line.startswith('|') and '|' in line[1:]:
                if not in_table:
                    in_table = True
                    # 解析表头
                    headers = [cell.strip() for cell in line.split('|')[1:-1]]
                    ws.append(headers)
                else:
                    # 检查是否是分隔行
                    if line.replace(' ', '').replace('-', '').replace('|', '') == '':
                        continue
                    
                    # 解析数据行
                    row_data = [cell.strip() for cell in line.split('|')[1:-1]]
                    if len(row_data) == len(headers):
                        ws.append(row_data)
            
            # 检测表格结束
            elif in_table and not line.startswith('|'):
                in_table = False
                headers = []
                # 添加空行分隔
                ws.append([])
        
        # 自动调整列宽
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(str(output_path))
        return output_path